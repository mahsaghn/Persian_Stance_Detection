# -*- coding: utf-8 -*-
"""PSFeatureExtractor.ipynb

Automatically generated by Colaboratory.

Original file is located at
    https://colab.research.google.com/drive/1-zwppOfr0Cr1k15U5cOBA8Kd3xdKSXWr
"""

# -*- coding: utf-8 -*-
import nltk
import string

import pandas
import pandas as pd
import re
import os
import numpy as np
import joblib
import os.path
import warnings
import stanza

from hazm import *
from hazm import word_tokenize as hazm_word_tokenize
from sklearn.feature_extraction.text import TfidfVectorizer
from difflib import SequenceMatcher
from nltk import word_tokenize as nltk_word_tokenize
from openpyxl import load_workbook
from sklearn.feature_extraction.text import CountVectorizer
from config.baseline_h2c import FeatureExtractorConf
from transformers import AutoConfig, AutoTokenizer, TFAutoModel

"""BaseLineWithGridSearch.ipynb

Automatically generated by Colaboratory.

Original file is located at
    https://colab.research.google.com/drive/1cxXw92aQZKvWJGmOdKIYGGXRMpK_XNvQ

# Import Required Libraries
"""

warnings.filterwarnings('ignore')

nltk.download('punkt')
stanza.download('fa')
with open('dataset/refute_words.txt', 'r') as refute_file:
    refute_hedge_reporte_words = [w.replace('\n', '') for w in refute_file.readlines()]


class PSFeatureExtractor():

    def __init__(self, cfg: FeatureExtractorConf, load_data=False):
        self.cfg = cfg
        self.important_words = ['؟', 'تکذیب', 'تکذیب شد', ':']
        self.clean_claims_headlines = []
        self.clean_claims = []
        self.clean_headlines = []
        self.fa_stop_words = self.__get_stop_words()
        self.fa_punctuations = self.__get_ponctuations()
        self.denied_words = self.fa_stop_words + list(string.punctuation) + list(self.fa_punctuations)

        if load_data:
            self.claims, self.headlines, self.isQuestion, self.hasTowParts, self.labels = self.__read_dataset()
        else:
            self.claims, self.headlines, self.isQuestion, self.hasTowParts, self.labels = self.__generate_dataset()

    def __get_ponctuations(self):
        with open(self.cfg.ponctuations_path) as f:
            poncs = [i.replace('\n','') for i in f.readlines()]
        return poncs

    def __get_stop_words(self):
        normalizer = Normalizer()
        with open(self.cfg.stopWord_path) as f:
            lineList = [normalizer.normalize(line.rstrip("\n\r")) for line in f]
        return lineList

    def __generate_dataset(self):
        data = pd.read_csv(self.cfg.dataset_path, encoding='utf-8')
        data = self.__remove_from_dataset(data)
        claims = np.array(data[self.cfg.claim_name].values)
        headlines = np.array(data[self.cfg.headline_name].values)
        labels = np.array(data[self.cfg.label_name].values)
        isQuestion = np.array([1 if '؟' in claim else 0 for claim in claims])
        hasTowParts = np.zeros(claims.shape[0])
        assert (claims.shape == headlines.shape == isQuestion.shape == labels.shape == hasTowParts.shape), "The features size are not equal."
        print('data shape is: ', claims.shape[0])
        return claims, headlines, isQuestion, hasTowParts, labels

    def __read_dataset(self):
        df = pd.read_csv(self.cfg.dataset_path, encoding='utf-8')
        df = self.__remove_from_dataset(df)
        claims = df[self.cfg.claim_name].values
        headlines = df[self.cfg.headline_name].values
        isQuestion = df[self.cfg.question_name].values
        hasTowParts = df[self.cfg.part_name].values
        labels = df[self.cfg.label_name].values
        assert (claims.shape == headlines.shape == isQuestion.shape == labels.shape == hasTowParts.shape), "The features size are not equal."
        print('data shape is: ',claims.shape)
        return claims, headlines, isQuestion, hasTowParts, labels

    def __remove_from_dataset(self, df):
        with open(self.cfg.uniq_claims_path) as f:
            uniq_claims = [i.replace('\n','') for i in f.readlines()]
        uni_number = 0
        if bool(uniq_claims):
            df['repeated'] = np.zeros(len(df), dtype=int)
            for claim_index in range(len(df[self.cfg.claim_name])):
                if (df[self.cfg.claim_name][claim_index] in uniq_claims) and uniq_claims[
                    df[self.cfg.claim_name][claim_index]] == df[self.cfg.headline_name][claim_index]:
                    df['repeated'][claim_index] = 1
                    print(df[self.cfg.claim_name][claim_index])
                    print(df[self.cfg.headline_name][claim_index])
                    uni_number += 1
            df.sort_values(by=['repeated'])
            print(df['repeated'])
        self.uniq_number = uni_number
        return df

    def clean_sentences(self):
        claims_result = []
        headlines_result = []
        parsbert_tokenizer = AutoTokenizer.from_pretrained(self.cfg.bert_model_path)
        for claim, headline in zip(self.claims, self.headlines):
            tokens = parsbert_tokenizer.tokenize(claim)
            clean_words = []
            for token in tokens:
                if token not in self.denied_words:
                    clean_words.append(token)
            new_sentence_c = parsbert_tokenizer.convert_tokens_to_string(clean_words)
            self.clean_claims.append(new_sentence_c)

            # headline
            tokens = parsbert_tokenizer.tokenize(headline)
            clean_words = []
            for token in tokens:
                if token not in self.denied_words:
                    clean_words.append(token)
            new_sentence_h = parsbert_tokenizer.convert_tokens_to_string(clean_words)
            self.clean_headlines.append(new_sentence_h)

            self.clean_claims_headlines.append(new_sentence_c + ' ' + new_sentence_h)

    def clean_sentence(self, sentence):
        normalizer = Normalizer()
        shayee = normalizer.normalize("شایعه")
        patterns = ["(/(\s)*" + shayee + "(\s)*[0-9]+)|(/(\s)*شایعه(\s)*[0-9]+)",
                    "/(\s)*[0-9]+",
                    "\\u200c|\\u200d|\\u200e|\\u200b|\\u2067|\\u2069"]
        clean_sentences = sentence
        for pattern in patterns:
            x = re.search(pattern, clean_sentences)
            if x:
                clean_sentences = re.sub(pattern, "", clean_sentences)
        punc_regex = re.compile('|'.join(map(re.escape, list(string.punctuation) + list(self.fa_punctuations))))
        clean_sentences = punc_regex.sub("", clean_sentences)
        return clean_sentences

    def clean_tokens(self, target_list):
        assert isinstance(target_list, (list)) == True, "Type of target_list is not correct. It has to be list."
        normalizer = Normalizer()
        clean_words = [i for item in target_list for i in item if normalizer.normalize(i) not in self.denied_words]
        return clean_words

    def stanford_tokenize(self, just_get_tokenized_words=False):

        nlp = stanza.Pipeline(lang='fa')
        claims_processors_result = []
        headlines_processors_result = []
        claims_tokenize = []
        headlines_tokenize = []

        for i in range(0, self.claims.shape[0]):
            clean_claim = self.clean_sentence(self.claims[i])
            self.clean_claims.append(clean_claim)
            doc = nlp(clean_claim)  # Run the pipeline on input text
            claims_processors_result.append(doc.sentences[0].words)
            claims_tokenize.append((obj.text for obj in doc.sentences[0].words))

            # headline
            clean_headline = self.clean_sentence(self.headlines[i])
            self.clean_headlines.append(clean_headline)
            doc = nlp(clean_headline)  # Run the pipeline on input text
            headlines_processors_result.append(doc.sentences[0].words)
            headlines_tokenize.append((obj.text for obj in doc.sentences[0].words))
            self.clean_claims_headlines.append(clean_claim + ' ' + clean_headline)

        self.tokens_claims = self.clean_tokens(target_list=claims_tokenize)
        self.tokens_headlines = self.clean_tokens(target_list=headlines_tokenize)
        if just_get_tokenized_words:
            return self.tokens_claims, self.tokens_headlines
        return claims_processors_result, headlines_processors_result

    def hazm_tokenize(self):
        claims_result = [self.clean_sentence(claim) for claim in self.claims]
        headlines_result = [self.clean_sentence(headline) for headline in self.headlines]
        self.clean_claims_headlines = [
            ' '.join(hazm_word_tokenize(claims_result[i])+hazm_word_tokenize(headlines_result[i])) for i in
            range(0, self.claims.shape[0])]
        self.tokens_claims = self.clean_tokens(target_list=[hazm_word_tokenize(claim_result) for claim_result in claims_result])
        self.tokens_headlines = self.clean_tokens(target_list=[hazm_word_tokenize(headline_result) for headline_result in headlines_result])
        return self.tokens_claims, self.tokens_headlines

    def nltk_tokenize(self):
        claims_result = [self.clean_sentence(claim) for claim in self.claims]
        headlines_result = [self.clean_sentence(headline) for headline in self.headlines]
        self.clean_claims_headlines = [
            ' '.join(nltk_word_tokenize(claims_result[i]) + nltk_word_tokenize(headlines_result[i])) for i in
            range(0, self.claims.shape[0])]
        self.tokens_claims = self.clean_tokens(target_list=[nltk_word_tokenize(claim_result) for claim_result in claims_result])
        self.tokens_headlines = self.clean_tokens( target_list=[nltk_word_tokenize(headline_result) for headline_result in headlines_result])
        return self.tokens_claims, self.tokens_headlines

    def tf_idf(self):
        tfidf = TfidfVectorizer(sublinear_tf=True, min_df=10, norm='l2', ngram_range=(1, 2))
        features = tfidf.fit_transform(self.clean_claims_headlines).toarray()
        return features

    def tf_idf(self):
        tfidf = TfidfVectorizer(sublinear_tf=True, min_df=10, norm='l2', ngram_range=(1, 2))
        features = tfidf.fit_transform(self.clean_claims_headlines).toarray()
        return features

    def similarity(self):
        feature = []
        for i, (claim, headline) in enumerate(zip(self.clean_claims, self.clean_headlines)):
            ratio = SequenceMatcher(None, claim, headline).ratio()
            quick_ratio = SequenceMatcher(None, claim, headline).quick_ratio()
            real_quick_ratio = SequenceMatcher(None, claim, headline).real_quick_ratio()
            feature.append([ratio, quick_ratio, real_quick_ratio])
        return feature

    def calc_important_words(self):
        assert (
                self.important_words != None), 'For calculating important words you should pass important words in initializer.'
        features = np.zeros((len(self.clean_claims_headlines), len(self.important_words)))
        for i in range(len(self.clean_claims_headlines)):
            for j in range(len(self.important_words)):
                if self.important_words[j] in self.clean_claims_headlines[i]:
                    features[i][j] = 1
        return features

    def calculate_root_distance(self, target_sentences=None):  # target_sentences = clean_headlines

        if target_sentences == None:
            target_sentences = self.clean_headlines

        nlp = stanfordnlp.Pipeline(lang='fa', models_dir=self.cfg.stanford_models_path, treebank=None, use_gpu=True)
        root_distance_feature = np.zeros((len(target_sentences), 1))
        for index, headline in enumerate(target_sentences):
            root_distance_feature[index] = -1
            doc = nlp(headline)
            root = [(i, doc.sentences[0].words[i].text) for i in range(len(doc.sentences[0].words)) if
                    doc.sentences[0].words[i].dependency_relation == 'root']
            if (len(root) == 0):
                continue

            root_index, root_word = root[0]

            for word_index, word in enumerate(headline.split()):
                target = [(i, refute_hedge_reporte_words[i]) for i in range(len(refute_hedge_reporte_words)) if
                          refute_hedge_reporte_words[i] == word]
                if len(target) > 0:
                    target_index, target_word = target[0]
                    root_distance_feature[index] = abs(word_index - root_index)
                    break
        return root_distance_feature

    def load_polarity_dataset(self):
        excel = load_workbook(filename=self.cfg.polarity_dataset_path)
        sheet = excel.active
        words_polarity_fa = {}
        for row in sheet.iter_rows():
            if row[2].value == "Polarity" or row[2].value == None:
                continue
            words_polarity_fa[row[0].value] = row[2].value
        return words_polarity_fa

    def calculate_polarity(self, target_sentences=None):
        words_polarity_fa = self.load_polarity_dataset()

        if target_sentences == None:
            target_sentences = zip(self.tokens_claims, self.tokens_headlines)

        # unzipping values
        mapped = list(target_sentences)
        claims, headlines = zip(*mapped)
        claims_array = np.asarray(claims)
        polarity_vector = np.zeros((len(claims_array), 30))

        for i, (claim, headline) in enumerate(zip(claims, headlines)):
            j = 0
            while j < len(claim) and j < 15:
                if claim[j] in words_polarity_fa:
                    polarity_vector[i][j] = words_polarity_fa[claim[j]]
                j += 1
            j = 0
            while j < len(headline) and j < 15:
                if headline[j] in words_polarity_fa:
                    polarity_vector[i][j + 15] = words_polarity_fa[headline[j]]
                j += 1
        return polarity_vector

    # Function to average all word vectors in a paragraph
    def __feature_vector_method(self, words, model, num_features):
        # Pre-initialising empty numpy array for speed
        featureVec = np.zeros(num_features, dtype="float32")
        nwords = 0
        # Converting Index2Word which is a list to a set for better speed in the execution.
        index2word_set = set(model.wv.index2word)

        for nwords, word in enumerate(words):
            if word in index2word_set:
                featureVec = np.add(featureVec, model[word])

        # Dividing the result by number of words to get average
        featureVec = np.divide(featureVec, nwords)
        return featureVec

    def get_w2v_feature(self, model, num_features, target_sentences=None):

        if target_sentences is None:
            target_sentences = self.clean_claims_headlines

        reviewFeatureVecs = np.zeros((len(target_sentences), num_features), dtype="float32")
        for counter, sentence in enumerate(target_sentences):
            # Printing a status message every 10000th review
            if counter % 1000 == 0:
                print("data %d of %d" % (counter, len(target_sentences)))

            reviewFeatureVecs[counter] = self.__feature_vector_method(sentence, model, num_features)

        return reviewFeatureVecs

    def get_bow(self, target_sentences=None):
        if target_sentences is None:
            target_sentences = self.clean_claims_headlines
        vectorizer = CountVectorizer(ngram_range=(1, 2))
        X = vectorizer.fit_transform(target_sentences)
        return X.toarray()

    def generate_Features(self):
        features = self.isQuestion
        features = np.reshape(features, (len(features), 1))
        file_name = ''

        if self.cfg.load_if_exist or self.cfg.save_feature:
            if self.cfg.tfidf:
                file_name += 'tfidf_'
            if self.cfg.similarity:
                file_name += 'similarity_'
            if self.cfg.important_words:
                file_name += 'important_words_'
            if self.cfg.is_question:
                file_name += 'is_question_'
            if self.cfg.more_than2_parts:
                file_name += 'more_than2_parts_'
            if self.cfg.root_distance:
                file_name += 'root_distance_'
            if self.cfg.polarity:
                file_name += 'polarity_'
            if self.cfg.w2v:
                file_name += 'w2v_'
            if self.cfg.bow:
                file_name += 'bow_'

        if self.cfg.load_if_exist:
            assert len(self.cfg.load_path) > 0, "Please enter load_path."
            load_file_name = self.cfg.load_path + '/' + file_name + '.pkl'
            if os.path.isfile(load_file_name):
                features = joblib.load(load_file_name)
                print('Features loaded successfully.')
                return features, file_name
            else:
                print('Features vector file is not exist.')
                # -------------- tfidf ----------
        if self.cfg.tfidf:
            print('Start to generate tf_idf feature')
            tf_idf_feature = self.tf_idf()
            features = np.append(features, tf_idf_feature, axis=1)
            print('End of tf_idf feature')
        # -------------- similarity ----------
        if self.cfg.similarity:
            print('Start to generate similarity feature')
            similarity_feature = self.similarity()
            features = np.append(features, similarity_feature, axis=1)
            print('End of similarity feature')
        # -------------- important words ----------
        if self.cfg.important_words:
            print('Start to generate important words feature')
            important_words_feature = self.calc_important_words()
            features = np.append(features, important_words_feature, axis=1)
            print('End of important words feature')
        # -------------- is question ----------
        if not self.cfg.is_question:
            features = features[:, 1:]
        else:
            print('"is question" feature was added.')
        # -------------- more than tow parts ----------
        if self.cfg.more_than2_parts:
            features = np.append(features, np.reshape(self.hasTowParts, (len(self.hasTowParts), 1)), axis=1)
            print('"more than tow parts" feature was added.')
        # -------------- root distance ----------
        if self.cfg.root_distance:
            print('Start to generate root distance feature')
            root_distance_feature = self.calculate_root_distance()
            features = np.append(features, root_distance_feature, axis=1)
            print('End of root distance feature')
        # -------------- root distance ----------
        if self.cfg.polarity:
            print('Start to generate polarity feature')
            polarity_feature = self.calculate_polarity()
            features = np.append(features, polarity_feature, axis=1)
            print('End of polarity feature')
            # -------------- w2v ----------
        if self.cfg.w2v:
            print('Start to generate w2v feature')
            assert len(self.cfg.w2v_model_path) > 0, "Please enter w2v_model_path."
            w2v_model = joblib.load(self.cfg.w2v_model_path)
            w2v_feature = self.get_w2v_feature(w2v_model, num_features=300)
            w2v_feature = (w2v_feature - np.min(w2v_feature)) / (np.max(w2v_feature) - np.min(w2v_feature))
            features = np.append(features, w2v_feature, axis=1)
            print('End of w2v feature')
            # -------------- bow ----------
        if self.cfg.bow:
            print('Start to generate bow feature')
            bow_feature = self.get_bow()
            features = np.append(features, bow_feature, axis=1)
            print('End of bow feature')

        if self.cfg.save_feature:
            joblib.dump(features, (self.cfg.save_path + '/' + file_name + '.pkl'))
            print('Features saved successfully.')
        return features, file_name
